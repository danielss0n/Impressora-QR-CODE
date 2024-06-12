from openpyxl.drawing.image import Image
from openpyxl import load_workbook

import tkinter as tk
from tkinter import Spinbox
from tkinter import messagebox 

from PIL import Image as PILImage
import qrcode

import win32com.client as win32
from datetime import date
import os

root = None  
ETIQUETA_FILE_NAME = "etiqueta.xlsx"
QR_IMAGE_PATH = "qrcode.png"
celula_dados = "A1"

def janela_inicial():
    global root 

    root = tk.Tk()
    root.geometry("400x300")
    root.iconbitmap("printer.ico")
    root.title("Impressora de etiquetas")
    fonte = ("Helvetica", 16)


    tk.Label(root, text="MATERIAL", font=fonte, justify="left", anchor="w").pack()
    material = tk.Entry(root, width=50, font=fonte)
    material.pack()

    tk.Label(root, text="PROJETO PEP", font=fonte, justify="left", anchor="w").pack()
    projeto = tk.Entry(root, width=50, font=fonte)
    projeto.pack()

    tk.Label(root, text="ORDEM", font=fonte, justify="left", anchor="w").pack()
    ordem = tk.Entry(root, width=50, font=fonte)
    ordem.pack()

    tk.Label(root, text="QUANTIDADE IMPRESSÕES", font=fonte, justify="left", anchor="w").pack()
    impressoes = tk.Spinbox(root, from_=0, to=100, font=fonte, textvariable=tk.DoubleVar(value=1))
    impressoes.pack()
    

    tk.Button(root, text="IMPRIMIR", font=fonte, command=lambda: iniciar_processo(material, projeto, ordem, impressoes)).pack()
    root.mainloop()


def iniciar_processo(_material, _projeto, _ordem, _impressoes):
    quantia = int(_impressoes.get())
    material = _material.get()
    ordem = _ordem.get()
    projeto = _projeto.get()
    dados = [material, ordem, projeto]

    if quantia > 10 or quantia < 1:
        messagebox.showwarning(title="Erro", message="Quantia não pode ser menor que 1 ou passar de 10 impressões")
        root.destroy()
        exit()

    gerar_qr(material)
    salvar_qr_na_planilha(dados)
    imprimir_etiqueta_impressora(quantia)


def salvar_qr_na_planilha(dados):
    global root

    wb = load_workbook(ETIQUETA_FILE_NAME)
    ws = wb.active

    planilha_limpar_imagens(ws)
    planilha_adicionar_imagem(ws, QR_IMAGE_PATH, "C1")
    planilha_adicionar_dados(ws, dados)

    wb.save(ETIQUETA_FILE_NAME)
    root.destroy()

    
def planilha_limpar_imagens(ws):
    for img in ws._images:
        ws._images.remove(img)

def planilha_adicionar_imagem(ws, image_path, cell):
    file = PILImage.open(image_path)
    img = Image(file)
    ws.add_image(img, cell)

def planilha_adicionar_dados(ws, dados):
    ws['A1'] = dados[0]
    ws['A3'] = dados[1]
    ws['A4'] = dados[2]
    ws['A6'] = date.today().strftime('%d/%m/%Y')
    

def imprimir_etiqueta_impressora(impressoes):
    excel = win32.Dispatch("Excel.Application")
    diretorio_atual = os.path.abspath(os.getcwd())
    caminho_arquivo = os.path.join(diretorio_atual, "etiqueta.xlsx")
    workbook = excel.Workbooks.Open(caminho_arquivo)

    for _ in range(impressoes):
        workbook.PrintOut()

    workbook.Close(False)
    excel.Quit()


def gerar_qr(dados):
    qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=6, border=0)
    qr.add_data(dados)
    qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white")
    img.save(QR_IMAGE_PATH)

janela_inicial()